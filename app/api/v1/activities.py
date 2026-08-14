from datetime import date

from fastapi import APIRouter, Depends, HTTPException, Request, status
from fastapi.responses import StreamingResponse
import io
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.api.deps import get_current_user, get_db, require_role, HR_WRITE_ROLES, ADMIN_ROLES
from app.core.audit import log_audit
from app.api.deps import get_client_ip
from app.models.employee import Employee
from app.models.activity import ActivityCategory, ActivityLog, DailyActivityStatus
from app.models.user import User
from app.schemas.activity import ActivityCategoryResponse, ActivityLogCreate, ActivityLogUpdate, ActivityLogResponse, EmployeeSummary, DailyStatusUpdate, DailyStatusResponse

router = APIRouter(prefix="/activities", tags=["activities"])

@router.get("/categories", response_model=list[ActivityCategoryResponse])
def get_categories(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
) -> list[ActivityCategory]:
    categories = db.scalars(
        select(ActivityCategory).where(
            ActivityCategory.company_id == current_user.company_id,
            ActivityCategory.deleted_at.is_(None)
        )
    ).all()
    
    # Auto-seed basic categories if none exist for easy testing
    if not categories:
        seed_names = ["Deep Work / Focus", "Meetings / Calls", "Email / Comms", "Planning / Admin", "Collaboration", "Learning / Development", "Break"]
        for name in seed_names:
            cat = ActivityCategory(company_id=current_user.company_id, name=name)
            db.add(cat)
        db.commit()
        categories = db.scalars(
            select(ActivityCategory).where(
                ActivityCategory.company_id == current_user.company_id,
                ActivityCategory.deleted_at.is_(None)
            )
        ).all()
        
    return list(categories)


@router.post("/log", response_model=ActivityLogResponse, status_code=status.HTTP_201_CREATED)
def log_activity(
    payload: ActivityLogCreate,
    request: Request,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
) -> ActivityLog:
    employee = db.scalar(select(Employee).where(Employee.user_id == current_user.id, Employee.deleted_at.is_(None)))
    if employee is None:
        raise HTTPException(status_code=404, detail="Employee profile not found")

    category = db.scalar(select(ActivityCategory).where(ActivityCategory.id == payload.category_id, ActivityCategory.company_id == current_user.company_id))
    if category is None:
        raise HTTPException(status_code=400, detail="Invalid category")

    status_entry = db.scalar(
        select(DailyActivityStatus).where(
            DailyActivityStatus.employee_id == employee.id,
            DailyActivityStatus.log_date == payload.log_date
        )
    )
    if status_entry and status_entry.status in ["On Leave", "Holiday"]:
        raise HTTPException(status_code=400, detail=f"Cannot log activities on a day marked as {status_entry.status}")

    # Check if a log already exists for this exact time block
    existing_log = db.scalar(
        select(ActivityLog).where(
            ActivityLog.employee_id == employee.id,
            ActivityLog.log_date == payload.log_date,
            ActivityLog.time_block == payload.time_block
        )
    )
    
    if existing_log:
        # Update existing
        existing_log.category_id = payload.category_id
        existing_log.duration_minutes = payload.duration_minutes
        existing_log.is_overtime = payload.is_overtime
        existing_log.notes = payload.notes
        db.commit()
        db.refresh(existing_log)
        
        # Attach category for response
        existing_log.category = category
        return existing_log
    else:
        # Create new
        new_log = ActivityLog(
            employee_id=employee.id,
            category_id=payload.category_id,
            log_date=payload.log_date,
            time_block=payload.time_block,
            duration_minutes=payload.duration_minutes,
            is_overtime=payload.is_overtime,
            notes=payload.notes
        )
        db.add(new_log)
        db.commit()
        db.refresh(new_log)
        
        # Attach category for response
        new_log.category = category
        
        log_audit(
            db,
            user_id=current_user.id,
            action="activity_logged",
            entity_type="activity_log",
            entity_id=new_log.id,
            ip_address=get_client_ip(request),
        )
        return new_log

@router.delete("/log/{log_id}", status_code=status.HTTP_204_NO_CONTENT)
def delete_activity(
    log_id: int,
    request: Request,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    employee = db.scalar(select(Employee).where(Employee.user_id == current_user.id, Employee.deleted_at.is_(None)))
    if employee is None:
        raise HTTPException(status_code=404, detail="Employee profile not found")

    log_entry = db.scalar(select(ActivityLog).where(ActivityLog.id == log_id, ActivityLog.employee_id == employee.id))
    if not log_entry:
        raise HTTPException(status_code=404, detail="Activity log not found")

    db.delete(log_entry)
    db.commit()

    log_audit(
        db,
        user_id=current_user.id,
        action="activity_deleted",
        entity_type="activity_log",
        entity_id=log_id,
        ip_address=get_client_ip(request),
    )

@router.put("/log/{log_id}", response_model=ActivityLogResponse)
def update_activity(
    log_id: int,
    payload: ActivityLogUpdate,
    request: Request,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    employee = db.scalar(select(Employee).where(Employee.user_id == current_user.id, Employee.deleted_at.is_(None)))
    if employee is None:
        raise HTTPException(status_code=404, detail="Employee profile not found")

    log_entry = db.scalar(select(ActivityLog).where(ActivityLog.id == log_id, ActivityLog.employee_id == employee.id))
    if not log_entry:
        raise HTTPException(status_code=404, detail="Activity log not found")

    category = db.scalar(select(ActivityCategory).where(ActivityCategory.id == payload.category_id, ActivityCategory.company_id == current_user.company_id))
    if category is None:
        raise HTTPException(status_code=400, detail="Invalid category")

    status_entry = db.scalar(
        select(DailyActivityStatus).where(
            DailyActivityStatus.employee_id == employee.id,
            DailyActivityStatus.log_date == payload.log_date
        )
    )
    if status_entry and status_entry.status in ["On Leave", "Holiday"]:
        raise HTTPException(status_code=400, detail=f"Cannot log activities on a day marked as {status_entry.status}")

    log_entry.category_id = payload.category_id
    log_entry.log_date = payload.log_date
    log_entry.time_block = payload.time_block
    log_entry.duration_minutes = payload.duration_minutes
    log_entry.is_overtime = payload.is_overtime
    log_entry.notes = payload.notes

    db.commit()
    db.refresh(log_entry)
    log_entry.category = category

    log_audit(
        db,
        user_id=current_user.id,
        action="activity_updated",
        entity_type="activity_log",
        entity_id=log_entry.id,
        ip_address=get_client_ip(request),
    )
    return log_entry


@router.get("/me", response_model=list[ActivityLogResponse])
def my_activities(
    start_date: date | None = None,
    end_date: date | None = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    employee = db.scalar(select(Employee).where(Employee.user_id == current_user.id, Employee.deleted_at.is_(None)))
    if employee is None:
        return []

    query = select(ActivityLog, ActivityCategory).join(ActivityCategory).where(ActivityLog.employee_id == employee.id)
    
    if start_date:
        query = query.where(ActivityLog.log_date >= start_date)
    if end_date:
        query = query.where(ActivityLog.log_date <= end_date)
        
    query = query.order_by(ActivityLog.log_date, ActivityLog.time_block)
    
    results = db.execute(query).all()
    
    # Map for response
    response_logs = []
    for log, cat in results:
        log.category = cat
        response_logs.append(log)
    return response_logs


@router.get("/team-summary", response_model=list[EmployeeSummary])
def get_team_summary(
    start_date: date,
    end_date: date,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    manager = db.scalar(select(Employee).where(Employee.user_id == current_user.id, Employee.deleted_at.is_(None)))
    if manager is None:
        return []

    from app.models.team import TeamMapping
    mapped_employee_ids = db.scalars(
        select(TeamMapping.employee_id).where(TeamMapping.manager_id == manager.id)
    ).all()
    
    if not mapped_employee_ids:
        return []

    mapped_employees = db.execute(
        select(Employee.id, User.full_name)
        .join(User, Employee.user_id == User.id)
        .where(Employee.id.in_(mapped_employee_ids))
    ).all()

    summary_map = {}
    for emp_id, emp_name in mapped_employees:
        summary_map[emp_id] = {
            "employee_id": emp_id,
            "employee_name": emp_name,
            "total_regular_hours": 0.0,
            "total_overtime_hours": 0.0,
            "categories": {},
            "statuses": {}
        }

    logs = db.execute(
        select(ActivityLog, ActivityCategory, Employee.id)
        .join(ActivityCategory, ActivityLog.category_id == ActivityCategory.id)
        .join(Employee, ActivityLog.employee_id == Employee.id)
        .where(
            ActivityLog.employee_id.in_(mapped_employee_ids),
            ActivityLog.log_date >= start_date,
            ActivityLog.log_date <= end_date
        )
    ).all()

    for log, cat, emp_id in logs:
        if emp_id not in summary_map:
            continue
            
        hours = log.duration_minutes / 60.0
        if log.is_overtime:
            summary_map[emp_id]["total_overtime_hours"] += hours
        else:
            summary_map[emp_id]["total_regular_hours"] += hours
            
        if cat.name not in summary_map[emp_id]["categories"]:
            summary_map[emp_id]["categories"][cat.name] = 0.0
        summary_map[emp_id]["categories"][cat.name] += hours

    statuses = db.execute(
        select(DailyActivityStatus.employee_id, DailyActivityStatus.status)
        .where(
            DailyActivityStatus.employee_id.in_(mapped_employee_ids),
            DailyActivityStatus.log_date >= start_date,
            DailyActivityStatus.log_date <= end_date
        )
    ).all()

    for emp_id, status in statuses:
        if emp_id in summary_map:
            if status not in summary_map[emp_id]["statuses"]:
                summary_map[emp_id]["statuses"][status] = 0
            summary_map[emp_id]["statuses"][status] += 1

    response = []
    for emp_data in summary_map.values():
        from app.schemas.activity import CategoryHours
        cats = [CategoryHours(category_name=k, hours=round(v, 2)) for k, v in emp_data["categories"].items()]
        response.append(
            EmployeeSummary(
                employee_id=emp_data["employee_id"],
                employee_name=emp_data["employee_name"],
                total_regular_hours=round(emp_data["total_regular_hours"], 2),
                total_overtime_hours=round(emp_data["total_overtime_hours"], 2),
                categories=cats,
                statuses=emp_data["statuses"]
            )
        )
        
    return response

@router.get("/company-summary", response_model=list[EmployeeSummary])
def get_company_summary(
    start_date: date,
    end_date: date,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_role(*ADMIN_ROLES))
):
    company_employees = db.execute(
        select(Employee.id, User.full_name)
        .join(User, Employee.user_id == User.id)
        .where(Employee.company_id == current_user.company_id)
    ).all()
    
    summary_map = {}
    for emp_id, emp_name in company_employees:
        summary_map[emp_id] = {
            "employee_id": emp_id,
            "employee_name": emp_name,
            "total_regular_hours": 0.0,
            "total_overtime_hours": 0.0,
            "categories": {},
            "statuses": {}
        }

    logs = db.execute(
        select(ActivityLog, ActivityCategory, Employee.id)
        .join(ActivityCategory, ActivityLog.category_id == ActivityCategory.id)
        .join(Employee, ActivityLog.employee_id == Employee.id)
        .where(
            Employee.company_id == current_user.company_id,
            ActivityLog.log_date >= start_date,
            ActivityLog.log_date <= end_date
        )
    ).all()

    for log, cat, emp_id in logs:
        if emp_id not in summary_map:
            continue
            
        hours = log.duration_minutes / 60.0
        if log.is_overtime:
            summary_map[emp_id]["total_overtime_hours"] += hours
        else:
            summary_map[emp_id]["total_regular_hours"] += hours
            
        if cat.name not in summary_map[emp_id]["categories"]:
            summary_map[emp_id]["categories"][cat.name] = 0.0
        summary_map[emp_id]["categories"][cat.name] += hours

    statuses = db.execute(
        select(DailyActivityStatus.employee_id, DailyActivityStatus.status)
        .join(Employee, DailyActivityStatus.employee_id == Employee.id)
        .where(
            Employee.company_id == current_user.company_id,
            DailyActivityStatus.log_date >= start_date,
            DailyActivityStatus.log_date <= end_date
        )
    ).all()

    for emp_id, status in statuses:
        if emp_id in summary_map:
            if status not in summary_map[emp_id]["statuses"]:
                summary_map[emp_id]["statuses"][status] = 0
            summary_map[emp_id]["statuses"][status] += 1

    response = []
    for emp_data in summary_map.values():
        from app.schemas.activity import CategoryHours
        cats = [CategoryHours(category_name=k, hours=round(v, 2)) for k, v in emp_data["categories"].items()]
        response.append(
            EmployeeSummary(
                employee_id=emp_data["employee_id"],
                employee_name=emp_data["employee_name"],
                total_regular_hours=round(emp_data["total_regular_hours"], 2),
                total_overtime_hours=round(emp_data["total_overtime_hours"], 2),
                categories=cats,
                statuses=emp_data["statuses"]
            )
        )
        
    return response

def _build_multi_sheet_activity_workbook(
    employee_name: str,
    employee_code: str | None,
    logs: list,
    status_map: dict,
    start_date: date,
    end_date: date
) -> io.BytesIO:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from datetime import timedelta

    wb = openpyxl.Workbook()
    ws_summary = wb.active
    ws_summary.title = "Weekly Summary"

    navy_fill = PatternFill(start_color="1E1B4B", end_color="1E1B4B", fill_type="solid")
    indigo_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
    sub_header_fill = PatternFill(start_color="E2E8F0", end_color="E2E8F0", fill_type="solid")
    green_fill = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid")
    yellow_fill = PatternFill(start_color="FEF9C3", end_color="FEF9C3", fill_type="solid")

    font_title = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    font_header = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    font_bold = Font(name="Calibri", size=11, bold=True)
    font_regular = Font(name="Calibri", size=11)
    font_muted = Font(name="Calibri", size=10, italic=True, color="64748B")

    thin_border = Border(
        left=Side(style="thin", color="CBD5E1"),
        right=Side(style="thin", color="CBD5E1"),
        top=Side(style="thin", color="CBD5E1"),
        bottom=Side(style="thin", color="CBD5E1")
    )

    logs_by_date = {}
    for log, cat in logs:
        if log.log_date not in logs_by_date:
            logs_by_date[log.log_date] = []
        logs_by_date[log.log_date].append((log, cat))

    # --- 1. WEEKLY SUMMARY SHEET ---
    ws_summary.merge_cells("A1:G1")
    ws_summary["A1"] = "EMPLOYEE DAILY ACTIVITY TRACKER — WEEKLY REPORT"
    ws_summary["A1"].font = font_title
    ws_summary["A1"].fill = navy_fill
    ws_summary["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws_summary.row_dimensions[1].height = 35

    ws_summary.append([])
    ws_summary.append(["Employee Name:", employee_name, "", "Period:", f"{start_date.isoformat()} to {end_date.isoformat()}"])
    ws_summary.append(["Employee Code:", employee_code or "N/A", "", "Generated On:", date.today().isoformat()])
    ws_summary.append([])

    for r in [3, 4]:
        for c in [1, 4]:
            ws_summary.cell(row=r, column=c).font = font_bold

    summary_headers = ["Date", "Day Name", "Working Status", "Regular (Hrs)", "Overtime (Hrs)", "Total Hours", "Entries Logged"]
    ws_summary.append(summary_headers)
    hdr_row = ws_summary.max_row
    ws_summary.row_dimensions[hdr_row].height = 24

    for c_idx, h in enumerate(summary_headers, start=1):
        cell = ws_summary.cell(row=hdr_row, column=c_idx)
        cell.fill = indigo_fill
        cell.font = font_header
        cell.alignment = Alignment(horizontal="center", vertical="center")

    tot_reg_hrs = 0.0
    tot_ot_hrs = 0.0
    tot_entries = 0
    cat_summary = {}

    current_date = start_date
    while current_date <= end_date:
        d_logs = logs_by_date.get(current_date, [])
        d_status = status_map.get(current_date, "Working")

        reg_m = sum(l.duration_minutes for l, _ in d_logs if not l.is_overtime)
        ot_m = sum(l.duration_minutes for l, _ in d_logs if l.is_overtime)

        reg_h = round(reg_m / 60.0, 2)
        ot_h = round(ot_m / 60.0, 2)
        tot_h = round(reg_h + ot_h, 2)
        cnt = len(d_logs)

        tot_reg_hrs += reg_h
        tot_ot_hrs += ot_h
        tot_entries += cnt

        for log, cat in d_logs:
            c_name = cat.name if cat else "Other"
            cat_summary[c_name] = cat_summary.get(c_name, 0.0) + (log.duration_minutes / 60.0)

        ws_summary.append([
            current_date.isoformat(),
            current_date.strftime("%A"),
            d_status,
            reg_h,
            ot_h,
            tot_h,
            cnt
        ])
        r_idx = ws_summary.max_row
        for c_idx in range(1, 8):
            cell = ws_summary.cell(row=r_idx, column=c_idx)
            cell.border = thin_border
            cell.font = font_regular
            if c_idx in (4, 5, 6, 7):
                cell.alignment = Alignment(horizontal="right")

        current_date += timedelta(days=1)

    # Total Summary Row
    ws_summary.append([
        "TOTAL",
        "",
        "",
        round(tot_reg_hrs, 2),
        round(tot_ot_hrs, 2),
        round(tot_reg_hrs + tot_ot_hrs, 2),
        tot_entries
    ])
    tot_row = ws_summary.max_row
    ws_summary.row_dimensions[tot_row].height = 22
    for c_idx in range(1, 8):
        cell = ws_summary.cell(row=tot_row, column=c_idx)
        cell.font = font_bold
        cell.fill = sub_header_fill
        cell.border = thin_border
        if c_idx in (4, 5, 6, 7):
            cell.alignment = Alignment(horizontal="right")

    # Category summary table
    ws_summary.append([])
    ws_summary.append(["Category Breakdown", "Total Hours Logged", "% of Total Hours"])
    cat_hdr_row = ws_summary.max_row
    ws_summary.row_dimensions[cat_hdr_row].height = 22
    for c_idx in range(1, 4):
        cell = ws_summary.cell(row=cat_hdr_row, column=c_idx)
        cell.fill = indigo_fill
        cell.font = font_header
        cell.alignment = Alignment(horizontal="center", vertical="center")

    all_hrs = tot_reg_hrs + tot_ot_hrs
    for c_name, c_hrs in cat_summary.items():
        pct = (c_hrs / all_hrs * 100.0) if all_hrs > 0 else 0.0
        ws_summary.append([c_name, round(c_hrs, 2), f"{pct:.1f}%"])
        r_idx = ws_summary.max_row
        for c_idx in range(1, 4):
            cell = ws_summary.cell(row=r_idx, column=c_idx)
            cell.border = thin_border
            cell.font = font_regular
            if c_idx in (2, 3):
                cell.alignment = Alignment(horizontal="right")

    for col in ws_summary.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws_summary.column_dimensions[col_letter].width = max(max_len + 3, 14)

    # --- 2. INDIVIDUAL DAY SHEETS ---
    current_date = start_date
    while current_date <= end_date:
        tab_name = current_date.strftime("%a %d-%b")  # e.g. "Mon 10-Aug"
        ws_day = wb.create_sheet(title=tab_name[:31])

        d_logs = logs_by_date.get(current_date, [])
        d_status = status_map.get(current_date, "Working")

        # Day Title Banner
        ws_day.merge_cells("A1:F1")
        ws_day["A1"] = f"DAILY ACTIVITY LOG — {current_date.strftime('%A, %B %d, %Y')}"
        ws_day["A1"].font = font_title
        ws_day["A1"].fill = navy_fill
        ws_day["A1"].alignment = Alignment(horizontal="center", vertical="center")
        ws_day.row_dimensions[1].height = 32

        ws_day.append([])
        ws_day.append(["Employee:", employee_name, "", "Working Status:", d_status])
        ws_day.append(["Date:", current_date.isoformat(), "", "Total Tasks Logged:", len(d_logs)])
        ws_day.append([])

        for r in [3, 4]:
            for c in [1, 4]:
                ws_day.cell(row=r, column=c).font = font_bold

        day_headers = ["Time Block", "Category", "Duration (Mins)", "Duration (Hrs)", "Type", "Notes / Task Details"]
        ws_day.append(day_headers)
        d_hdr_row = ws_day.max_row
        ws_day.row_dimensions[d_hdr_row].height = 24

        for c_idx, h in enumerate(day_headers, start=1):
            cell = ws_day.cell(row=d_hdr_row, column=c_idx)
            cell.fill = indigo_fill
            cell.font = font_header
            cell.alignment = Alignment(horizontal="center", vertical="center")

        d_reg_m = 0
        d_ot_m = 0

        if not d_logs:
            ws_day.append(["-", "No activities logged on this date", 0, 0.0, "-", ""])
            r_idx = ws_day.max_row
            for c_idx in range(1, 7):
                cell = ws_day.cell(row=r_idx, column=c_idx)
                cell.font = font_muted
                cell.border = thin_border
        else:
            for log, cat in d_logs:
                tb = log.time_block.strftime("%H:%M") if hasattr(log.time_block, "strftime") else str(log.time_block)[:5]
                mins = log.duration_minutes
                hrs = round(mins / 60.0, 2)
                type_str = "Overtime" if log.is_overtime else "Regular"

                if log.is_overtime:
                    d_ot_m += mins
                else:
                    d_reg_m += mins

                ws_day.append([
                    tb,
                    cat.name if cat else "Other",
                    mins,
                    hrs,
                    type_str,
                    log.notes or ""
                ])
                r_idx = ws_day.max_row
                for c_idx in range(1, 7):
                    cell = ws_day.cell(row=r_idx, column=c_idx)
                    cell.font = font_regular
                    cell.border = thin_border
                    if c_idx in (3, 4):
                        cell.alignment = Alignment(horizontal="right")
                    if c_idx == 5:
                        cell.alignment = Alignment(horizontal="center")
                        cell.fill = yellow_fill if log.is_overtime else green_fill

        # Day Summary Total
        d_tot_m = d_reg_m + d_ot_m
        ws_day.append([
            "DAILY TOTAL",
            f"Regular: {round(d_reg_m/60.0, 2)}h | Overtime: {round(d_ot_m/60.0, 2)}h",
            d_tot_m,
            round(d_tot_m / 60.0, 2),
            "",
            ""
        ])
        d_tot_row = ws_day.max_row
        ws_day.row_dimensions[d_tot_row].height = 22
        for c_idx in range(1, 7):
            cell = ws_day.cell(row=d_tot_row, column=c_idx)
            cell.font = font_bold
            cell.fill = sub_header_fill
            cell.border = thin_border
            if c_idx in (3, 4):
                cell.alignment = Alignment(horizontal="right")

        for col in ws_day.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = get_column_letter(col[0].column)
            ws_day.column_dimensions[col_letter].width = max(max_len + 3, 14)

        current_date += timedelta(days=1)

    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)
    return stream


@router.get("/export/me")
def export_my_activities(
    start_date: date,
    end_date: date,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    employee = db.scalar(select(Employee).where(Employee.user_id == current_user.id, Employee.deleted_at.is_(None)))
    if not employee:
        raise HTTPException(status_code=404, detail="Employee profile not found")

    logs = db.execute(
        select(ActivityLog, ActivityCategory)
        .join(ActivityCategory)
        .where(
            ActivityLog.employee_id == employee.id,
            ActivityLog.log_date >= start_date,
            ActivityLog.log_date <= end_date
        )
        .order_by(ActivityLog.log_date, ActivityLog.time_block)
    ).all()

    statuses = db.execute(
        select(DailyActivityStatus.log_date, DailyActivityStatus.status)
        .where(
            DailyActivityStatus.employee_id == employee.id,
            DailyActivityStatus.log_date >= start_date,
            DailyActivityStatus.log_date <= end_date
        )
    ).all()
    status_map = {row.log_date: row.status for row in statuses}

    emp_name = current_user.full_name or current_user.email
    stream = _build_multi_sheet_activity_workbook(emp_name, employee.employee_code, logs, status_map, start_date, end_date)

    filename = f"Activity_Report_{emp_name.replace(' ', '_')}_{start_date}_to_{end_date}.xlsx"
    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@router.get("/export/{employee_id}")
def export_employee_activities(
    employee_id: int,
    start_date: date,
    end_date: date,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_role(*ADMIN_ROLES))
):
    employee = db.scalar(select(Employee).where(Employee.id == employee_id, Employee.company_id == current_user.company_id))
    if not employee:
        raise HTTPException(status_code=404, detail="Employee not found")
        
    emp_user = db.get(User, employee.user_id)
    emp_name = emp_user.full_name if emp_user else f"Employee_{employee_id}"
    
    logs = db.execute(
        select(ActivityLog, ActivityCategory)
        .join(ActivityCategory)
        .where(
            ActivityLog.employee_id == employee_id,
            ActivityLog.log_date >= start_date,
            ActivityLog.log_date <= end_date
        )
        .order_by(ActivityLog.log_date, ActivityLog.time_block)
    ).all()
    
    statuses = db.execute(
        select(DailyActivityStatus.log_date, DailyActivityStatus.status)
        .where(
            DailyActivityStatus.employee_id == employee_id,
            DailyActivityStatus.log_date >= start_date,
            DailyActivityStatus.log_date <= end_date
        )
    ).all()
    status_map = {row.log_date: row.status for row in statuses}

    stream = _build_multi_sheet_activity_workbook(emp_name, employee.employee_code, logs, status_map, start_date, end_date)
    
    filename = f"Activities_{emp_name.replace(' ', '_')}_{start_date}_to_{end_date}.xlsx"
    return StreamingResponse(
        stream, 
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

@router.get("/status/{log_date}", response_model=DailyStatusResponse)
def get_daily_status(
    log_date: date,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    employee = db.scalar(select(Employee).where(Employee.user_id == current_user.id, Employee.deleted_at.is_(None)))
    if employee is None:
        raise HTTPException(status_code=404, detail="Employee profile not found")

    status_entry = db.scalar(
        select(DailyActivityStatus).where(
            DailyActivityStatus.employee_id == employee.id,
            DailyActivityStatus.log_date == log_date
        )
    )

    if not status_entry:
        return DailyActivityStatus(
            id=0,
            employee_id=employee.id,
            log_date=log_date,
            status="Working",
            notes=None
        )

    return status_entry


@router.put("/status/{log_date}", response_model=DailyStatusResponse)
def update_daily_status(
    log_date: date,
    payload: DailyStatusUpdate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    employee = db.scalar(select(Employee).where(Employee.user_id == current_user.id, Employee.deleted_at.is_(None)))
    if employee is None:
        raise HTTPException(status_code=404, detail="Employee profile not found")

    status_entry = db.scalar(
        select(DailyActivityStatus).where(
            DailyActivityStatus.employee_id == employee.id,
            DailyActivityStatus.log_date == log_date
        )
    )

    if status_entry:
        status_entry.status = payload.status
        status_entry.notes = payload.notes
    else:
        status_entry = DailyActivityStatus(
            employee_id=employee.id,
            log_date=log_date,
            status=payload.status,
            notes=payload.notes
        )
        db.add(status_entry)

    db.commit()
    db.refresh(status_entry)
    
    return status_entry
